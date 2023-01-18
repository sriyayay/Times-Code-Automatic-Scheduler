import openpyxl
from openpyxl import Workbook

dataframe = openpyxl.load_workbook('Schedule.xlsx')
dataframe1 = dataframe.active

#Creating a sheet within the workbook (this is the output excel file)
wb = Workbook()
wb.save("FinalSchedule.xlsx")
sheet1 = wb.create_sheet(index = 0, title = "Week Schedule")
wb.save("FinalSchedule.xlsx")

#Writing the Time Slots in the Ouput Excel Sheet
sheet1.cell(1, 1).value = "Times"
sheet1.cell(1, 2).value = "Person on Shift"
sheet1.cell(2, 1).value = "9:00"
sheet1.cell(3, 1).value = "10:00"
sheet1.cell(4, 1).value = "11:00"
sheet1.cell(5, 1).value = "12:00"
sheet1.cell(6, 1).value = "1:00"
sheet1.cell(7, 1).value = "2:00"
sheet1.cell(8, 1).value = "3:00"
sheet1.cell(9, 1).value = "4:00"
sheet1.cell(10, 1).value = "5:00"


#Reads input excel file, and writes name of person with that available time
#This code does not check for overlaps in available times
for i in range(2, dataframe1.max_row + 1):
  for j in range(3, dataframe1.max_column + 1):
   
    stg = str(dataframe1.cell(i, j).value)
    list = stg.split("-")
    time_output = 2
   
    for x in list:
      time_output = 2
      if x == "9":
        sheet1.cell(time_output, 2).value = dataframe1.cell(i, 1).value
      time_output += 1
      if x == "10":
        sheet1.cell(time_output, 2).value = dataframe1.cell(i, 1).value
      time_output += 1
      if x == "11":
        sheet1.cell(time_output, 2).value = dataframe1.cell(i, 1).value
      time_output += 1
      if x == "12":
        sheet1.cell(time_output, 2).value = dataframe1.cell(i, 1).value
      time_output += 1
      if x == "1":
        sheet1.cell(time_output, 2).value = dataframe1.cell(i, 1).value
      time_output += 1
      if x == "2":
        sheet1.cell(time_output, 2).value = dataframe1.cell(i, 1).value 
      time_output += 1
      if x == "3":
        sheet1.cell(time_output, 2).value = dataframe1.cell(i, 1).value        
      time_output += 1
      if x == "4":
        sheet1.cell(time_output, 2).value = dataframe1.cell(i, 1).value 
      time_output += 1
      if x == "5":
        sheet1.cell(time_output, 2).value = dataframe1.cell(i, 1).value    
wb.save('FinalSchedule.xlsx')

